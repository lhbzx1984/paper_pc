from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import uvicorn
import yaml
import io
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from services.document_parser import DocumentParser
from services.rag_evaluator import RAGEvaluator
from services.llm_client import LLMClient

app = FastAPI(title="论文评审系统")

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 加载配置
with open("config.yaml", "r", encoding="utf-8") as f:
    config = yaml.safe_load(f)

# 环境变量覆盖（生产环境优先使用环境变量，避免在config.yaml中暴露API Key）
if os.environ.get("LLM_API_KEY"):
    config["llm"]["api_key"] = os.environ["LLM_API_KEY"]
if os.environ.get("LLM_BASE_URL"):
    config["llm"]["base_url"] = os.environ["LLM_BASE_URL"]
if os.environ.get("LLM_MODEL"):
    config["llm"]["model"] = os.environ["LLM_MODEL"]

# 初始化服务
llm_client = LLMClient(config["llm"])
rag_evaluator = RAGEvaluator(config["evaluation"], llm_client)
document_parser = DocumentParser()


class ConfigUpdate(BaseModel):
    llm_api_key: Optional[str] = None
    llm_base_url: Optional[str] = None
    llm_model: Optional[str] = None


@app.get("/")
async def root():
    return {"message": "论文评审系统API"}


@app.post("/api/upload")
async def upload_paper(file: UploadFile = File(...)):
    """上传论文并进行评审"""
    try:
        # 检查文件类型
        allowed_extensions = [".docx", ".doc", ".pdf", ".wps"]
        file_ext = Path(file.filename).suffix.lower()
        
        if file_ext not in allowed_extensions:
            raise HTTPException(400, "不支持的文件格式")
        
        # 读取文件内容
        content = await file.read()
        
        # 解析文档
        text = document_parser.parse(content, file_ext)
        
        # 使用RAG进行评审
        result = await rag_evaluator.evaluate(text)
        
        return {
            "success": True,
            "filename": file.filename,
            "result": result
        }
    
    except Exception as e:
        raise HTTPException(500, f"评审失败: {str(e)}")


@app.get("/api/config")
async def get_config():
    """获取当前配置（隐藏敏感信息）"""
    return {
        "llm_model": config["llm"]["model"],
        "llm_base_url": config["llm"]["base_url"],
        "evaluation_criteria": config["evaluation"]["criteria"]
    }


@app.post("/api/config")
async def update_config(update: ConfigUpdate):
    """更新配置"""
    try:
        if update.llm_api_key:
            config["llm"]["api_key"] = update.llm_api_key
        if update.llm_base_url:
            config["llm"]["base_url"] = update.llm_base_url
        if update.llm_model:
            config["llm"]["model"] = update.llm_model
        
        # 保存配置
        with open("config.yaml", "w", encoding="utf-8") as f:
            yaml.dump(config, f, allow_unicode=True)
        
        # 重新初始化LLM客户端
        global llm_client, rag_evaluator
        llm_client = LLMClient(config["llm"])
        rag_evaluator = RAGEvaluator(config["evaluation"], llm_client)
        
        return {"success": True, "message": "配置更新成功"}
    
    except Exception as e:
        raise HTTPException(500, f"配置更新失败: {str(e)}")


@app.get("/api/criteria/export")
async def export_criteria():
    """导出当前评审标准为Excel模版"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评审标准"

    # 表头样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1890ff", end_color="1890ff", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["评审项名称", "权重", "满分", "评审说明"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 填入当前评审标准数据
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_idx, criterion in enumerate(config["evaluation"]["criteria"], 2):
        ws.cell(row=row_idx, column=1, value=criterion["name"]).alignment = data_align
        ws.cell(row=row_idx, column=2, value=criterion["weight"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3, value=criterion["max_score"]).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=4, value=criterion["description"]).alignment = data_align
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).border = thin_border

    # 列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60

    # 冻结表头
    ws.freeze_panes = "A2"

    # 写入内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=evaluation_criteria_template.xlsx"}
    )


@app.post("/api/criteria/import")
async def import_criteria(file: UploadFile = File(...)):
    """从Excel导入评审标准"""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # 读取表头，建立列映射
        header_row = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        col_map = {}
        header_aliases = {
            "name": ["评审项名称", "名称", "项目名称", "name"],
            "weight": ["权重", "weight"],
            "max_score": ["满分", "分值", "最高分", "max_score"],
            "description": ["评审说明", "说明", "描述", "标准说明", "description"],
        }
        for field, aliases in header_aliases.items():
            for col_idx, header in enumerate(header_row):
                if header.lower() in [a.lower() for a in aliases]:
                    col_map[field] = col_idx
                    break

        # 如果没匹配到表头，按默认列顺序处理
        if not col_map:
            col_map = {"name": 0, "weight": 1, "max_score": 2, "description": 3}

        # 解析数据行
        new_criteria = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # 跳过空行
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            name = str(row[col_map["name"]]).strip() if col_map.get("name") is not None and row[col_map["name"]] else ""
            if not name:
                errors.append(f"第{row_idx}行：评审项名称不能为空")
                continue

            try:
                weight = float(row[col_map["weight"]]) if col_map.get("weight") is not None else 0
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：权重格式错误（应为数字）")
                continue

            try:
                max_score = int(float(row[col_map["max_score"]])) if col_map.get("max_score") is not None else 0
            except (ValueError, TypeError):
                errors.append(f"第{row_idx}行：满分格式错误（应为整数）")
                continue

            description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None and row[col_map["description"]] else ""

            # 基础校验
            if weight < 0 or weight > 1:
                errors.append(f"第{row_idx}行：权重应在0-1之间（当前{weight}）")
                continue
            if max_score <= 0:
                errors.append(f"第{row_idx}行：满分应大于0（当前{max_score}）")
                continue

            new_criteria.append({
                "name": name,
                "weight": weight,
                "max_score": max_score,
                "description": description
            })

        if errors:
            raise HTTPException(400, f"导入校验失败：\n" + "\n".join(errors))

        if not new_criteria:
            raise HTTPException(400, "未读取到有效的评审标准数据，请检查Excel内容")

        # 权重总和校验（允许一定误差）
        total_weight = sum(c["weight"] for c in new_criteria)
        if abs(total_weight - 1.0) > 0.05:
            raise HTTPException(
                400,
                f"权重总和应为1.0，当前为{total_weight:.2f}，请调整后重新导入"
            )

        # 更新配置
        config["evaluation"]["criteria"] = new_criteria

        # 保存到config.yaml
        with open("config.yaml", "w", encoding="utf-8") as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False, sort_keys=True)

        # 热重载评审器
        global rag_evaluator
        rag_evaluator = RAGEvaluator(config["evaluation"], llm_client)

        return {
            "success": True,
            "message": f"成功导入{len(new_criteria)}项评审标准",
            "evaluation_criteria": new_criteria
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"导入失败: {str(e)}")


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
